import time
import sys
import subprocess
import serial
from serial.tools import list_ports
import os

import uuid
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
import re

def get_today_folder():
    today = datetime.now().strftime("%Y-%m-%d")
    base = Path("Gulliver Tests")
    folder = base / today
    folder.mkdir(parents=True, exist_ok=True)
    return folder, today

def log_to_excel(folder, date_str, row):
    xlsx = folder / f"{date_str}.xlsx"

    if xlsx.exists():
        wb = load_workbook(xlsx)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["TEST ID", "SN", "IMEI", "ICCID", "IMSI"])

    ws.append(row)
    wb.save(xlsx)

def extract_ids(log_lines):
    imei = iccid = imsi = "N/A"

    for line in log_lines:
        if "DEVICEINFO:" in line:
            m1 = re.search(r"IMEI:(\d+)", line)
            m2 = re.search(r"ICCID:(\d+)", line)
            m3 = re.search(r"IMSI:(\d+)", line)

            if m1: imei = m1.group(1)
            if m2: iccid = m2.group(1)
            if m3: imsi = m3.group(1)

    return imei, iccid, imsi

def write_txt(folder, imei, test_id, log_lines):
    fname = folder / f"{imei}.txt"

    with open(fname, "a", encoding="utf-8") as f:
        f.write(f"\nNew Test with ID: {test_id}\n")
        for line in log_lines:
            f.write(line + "\n")

def collect_test_log(ser, timeout=180):
    start = time.time()
    lines = []
    final_result = None

    while time.time() - start < timeout:
        line = ser.readline().decode(errors="ignore").strip()

        if not line:
            continue

        print("[DUT]", line)
        lines.append(line)

        if line.startswith("FINALRESULT:"):
            final_result = line.split(":",1)[1].strip()
            break

    return lines, final_result


ARDUINO_BAUD = 9600
ESP_BAUD = 921600

# ================= FLASH ARGS ESP32 =================
FLASH_ARGS = [
    "write_flash",
    "-z",
    "0x0000", r"C:\Users\DimitrisOikonomou\Desktop\Gulliver Testing\ESPFW\ESP32-C3-MINI-1-V3.2.0.0\bootloader\bootloader.bin",
    "0x60000", r"C:\Users\DimitrisOikonomou\Desktop\Gulliver Testing\ESPFW\ESP32-C3-MINI-1-V3.2.0.0\esp-at.bin",
    "0x8000", r"C:\Users\DimitrisOikonomou\Desktop\Gulliver Testing\ESPFW\ESP32-C3-MINI-1-V3.2.0.0\partition_table\partition-table.bin",
    "0xD000", r"C:\Users\DimitrisOikonomou\Desktop\Gulliver Testing\ESPFW\ESP32-C3-MINI-1-V3.2.0.0\ota_data_initial.bin",
    "0xF000", r"C:\Users\DimitrisOikonomou\Desktop\Gulliver Testing\ESPFW\ESP32-C3-MINI-1-V3.2.0.0\phy_multiple_init_data.bin",
    "0x1E000", r"C:\Users\DimitrisOikonomou\Desktop\Gulliver Testing\ESPFW\ESP32-C3-MINI-1-V3.2.0.0\at_customize.bin",
    "0x1F000", r"C:\Users\DimitrisOikonomou\Desktop\Gulliver Testing\ESPFW\ESP32-C3-MINI-1-V3.2.0.0\customized_partitions\mfg_nvs.bin",
]

# ================= PORT DETECTION =================
def detect_ports():
    arduino = None
    esp = None
    for p in list_ports.comports():
        desc = (p.description or "").lower()
        if "arduino" in desc or (p.vid == 0x2341 and p.pid == 0x1002):
            arduino = p.device
        elif any(x in desc.lower() for x in ["cp210", "ch340", "ftdi"]) or (p.vid == 0x0403 and p.pid == 0x6001):
            esp = p.device
    return arduino, esp

# ================= FLASH ESP =================
def run_flash(esp_port):
    cmd = [
        "python", "-m", "esptool",
        "--chip", "esp32c3",
        "--port", esp_port,
        "--baud", str(ESP_BAUD)
    ] + FLASH_ARGS
    print("[Test Jig] Flashing ESP:", " ".join(cmd))
    return subprocess.run(cmd).returncode == 0

# ================= FLASH MCU (J-Flash CLI)Manual =================
# def run_jflash_bin(jflash_path, bin_file, address, device):
#     cmd = [
#         jflash_path,
#         "/Program",
#         f"/File={bin_file}",
#         f"/Address={address}",
#         f"/Device={device}",
#         "/Autostart",
#         "/NoGUI",
#         "/Quit"
#     ]
#     print("[Test Jig] Running J-Flash Lite CLI:", " ".join(cmd))
#     result = subprocess.run(cmd)
#     return result.returncode == 0

# ================= FLASH MCU (J-Flash CLI)Auto Project =================
def run_jlink_flash(jlink_exe, command_file, log_file):
    os.makedirs(os.path.dirname(log_file), exist_ok=True)

    cmd = [
        jlink_exe,
        "-CommandFile", command_file,
        "-Log", log_file
    ]

    print("[Test Jig] Running J-Link Commander:")
    print(" ".join(cmd))

    result = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True
    )

    print(result.stdout)

    if result.returncode != 0:
        print("[ERR] J-Link returned:", result.returncode)
        return False

    if "Downloading file" in result.stdout or "O.K." in result.stdout:
        return True

    print("[ERR] J-Link did not report success")
    return False


# ================= MAIN =================
def main():
    arduino_port, esp_port = detect_ports()
    if not arduino_port or not esp_port:
        print("[ERR] Ports not found")
        sys.exit(1)

    print(f"[OK] Arduino @ {arduino_port}")
    print(f"[OK] ESP-TTL @ {esp_port}")

    with serial.Serial(arduino_port, ARDUINO_BAUD, timeout=5) as ser:
        time.sleep(2)
        ser.reset_input_buffer()
        ser.reset_output_buffer()

        # ================= HANDSHAKE =================
        ser.write(b'?\n')
        resp = ser.readline().decode(errors="ignore").strip()
        if resp != "ARDUINO_JIG_OK":
            print("[ERR] Arduino handshake failed")
            sys.exit(1)

        # ================= FLASH ESP FIRST =================
        print("[Test Jig] Sending 'P' to Arduino for ESP programming")
        ser.write(b'P\n')
        ser.flush()
        time.sleep(2.5) 

        print("[Test Jig] Flashing ESP...")
        if not run_flash(esp_port):
            print("[ERR] ESP Flash FAILED")
            sys.exit(1)

        # ser.write(b"R")  # release ESP
        # ser.flush()
        # time.sleep(0.5)

        # ================= FLASH MCU WITH J-LINK =================
        print("[Test Jig] Sending 'O' to Arduino to start MCU programming")
        ser.write(b'O\n')
        ser.flush()
        time.sleep(1.5)
        #Manual J-Flash CLI approach (works but is slow and less flexible than using a project file)
        # JFLASH_PATH = r"C:\Program Files\SEGGER\JLink_V866\JFlashLite.exe"
        # BIN_FILE = r"C:\Users\DimitrisOikonomou\Desktop\Gulliver Testing\Gulliver_Barista_19_jtag_Ryoma.bin"
        # ADDRESS = "0x0000"
        # DEVICE = "ATSAMD21J18A"
        # if not run_jflash_bin(JFLASH_PATH, BIN_FILE, ADDRESS, DEVICE):
        #     print("[ERR] J-Flash failed")
        #     sys.exit(1)

        #Auto J-Flash project approach (preferred for flexibility and ease of maintenance, just update the project file in J-Flash GUI and keep the script unchanged)
        JLINK_EXE = r"C:\Program Files\SEGGER\JLink_V866\JLink.exe"
        JLINK_SCRIPT = r"C:\Users\DimitrisOikonomou\Desktop\Gulliver Testing\flash.txt"
        JLINK_LOG = r"C:\Users\DimitrisOikonomou\Desktop\Gulliver Testing\logs\jlink_log.txt"

        if not run_jlink_flash(JLINK_EXE, JLINK_SCRIPT, JLINK_LOG):
            print("[ERR] MCU Flash FAILED")
            sys.exit(1)
        time.sleep(1.5)  # small delay before checking log
        if not run_jlink_flash(JLINK_EXE, JLINK_SCRIPT, JLINK_LOG):
            print("[ERR] MCU Flash FAILED")
            sys.exit(1)
        time.sleep(1.5)  # small delay before checking log

        # ============= START TESTING ==============
        print("[Test Jig] Sending 'T' to Arduino to start testing")
        ser.write(b'T\n')
        ser.flush()

        test_id = str(uuid.uuid4())

        log_lines, final_result = collect_test_log(ser, timeout=180)

        folder, date_str = get_today_folder()

        imei, iccid, imsi = extract_ids(log_lines)

        write_txt(folder, imei, test_id, log_lines)

        # ================= PASS / FAIL =================

        if final_result is None:
            print("❌ TEST TIMEOUT (3 min)")
            passed = False
        else:
            print(f"🏁 FINAL RESULT: {final_result}")
            passed = (final_result.upper() == "PASS")

        if passed:
            sn = input("Enter Serial Number (SN): ").strip()
            log_to_excel(folder, date_str, [test_id, sn, imei, iccid, imsi])
            print("✅ TEST PASSED")
        else:
            print("❌ TEST FAILED")

        # ============= FINAL HANDSHAKE ==============

        if not passed:
            ser.write(b"TESTFAILED\n")
            ser.flush()
            time.sleep(0.2)

        ser.write(b"TESTCOMPLETED\n")
        ser.flush()


    print("[Test Jig] ALL DONE ✅ Remove the PCBA from the jig.")
if __name__ == "__main__":
    main()
